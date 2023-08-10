from flask import Flask, render_template, Response, jsonify
import omronfins.finsudp as finsudp
from omronfins.finsudp import datadef
from openpyxl import load_workbook
from datetime import datetime
from queue import Queue
import queue
import threading
import datetime
import time

ruta='C:/Users/jorge.matus/Desktop/TESLA-srvr/Libro1.xlsx'
ruta2='C:/Users/jorge.matus/Desktop/TESLA-srvr/Libro2.xlsx'


# pa=valor de producción tca=tiempo ciclo auto tf=tiempo falla idalm= ID alarma Andon = Llamadas ANDON
pa=[0,0,0,0,0,0,0,0,0,0,0]
tca=[0,0,0,0,0,0,0,0,0,0,0]
tf=[0,0,0,0,0,0,0,0,0,0,0]
idalm=[0,0,0,0,0,0,0,0,0,0,0]
Andn=[0,0,0,0,0,0,0,0,0,0,0]
# AnMat=Llamada material AnLid=Llamada lider AnMtto=Llamada mantenimiento AnCal=Llamada Calidad 
status=[0,0,0,0,0,0,0,0,0,0,0]


#Ip string e INT, estaciones
Tesla=['RR100A','RR90A','RR80A','RR70A','RR60A','RR50A','FR80B','FR70B','FR60B','FR50B','FR40B']
ip=['36','35','34','33','32','31','25','24','23','22','21']
ipst=[36,35,34,33,32,31,25,24,23,22,21]

#Monitor de producción
#pl=plan de producción  df=diferencia plan vs actual  tt=tiempo total desde inicio de turno
#te = tiempo efectivo (min) th=tiempo hombre (diseño,segs)  pp=paro planeado (break, comedor)
#thp = tiempo hombre promedio vs producción (Min) tm=tiempo sin operación automatica
#tm = Tiempo muerto final min  tp= tiempo muerto por anomalía  tc=Tiempo ciclo de proceso
#ef= eficiencia calculada
pl=[0,0,0,0,0,0,0,0,0,0,0]
df=[0,0,0,0,0,0,0,0,0,0,0]
tt=[0,0,0,0,0,0,0,0,0,0,0]
te=[0,0,0,0,0,0,0,0,0,0,0]
th=[0,0,0,0,0,0,0,0,0,0,0]
pp=[0,0,0,0,0,0,0,0,0,0,0]
thp=[0,0,0,0,0,0,0,0,0,0,0]
tm=[0,0,0,0,0,0,0,0,0,0,0]
tmf=[0,0,0,0,0,0,0,0,0,0,0]
tp=[0,0,0,0,0,0,0,0,0,0,0]
tc=[0,0,0,0,0,0,0,0,0,0,0]
ef=[0,0,0,0,0,0,0,0,0,0,0]

#Monitor de ANDON
#mat=material ld=lea mt=mantenimiento  cal=calidad 
mat1=[0,0,0,0,0,0,0,0,0,0,0]
ld1=[0,0,0,0,0,0,0,0,0,0,0]
mt1=[0,0,0,0,0,0,0,0,0,0,0]
cal1=[0,0,0,0,0,0,0,0,0,0,0]
mat2=[0,0,0,0,0,0,0,0,0,0,0]
ld2=[0,0,0,0,0,0,0,0,0,0,0]
mt2=[0,0,0,0,0,0,0,0,0,0,0]
cal2=[0,0,0,0,0,0,0,0,0,0,0]

#Monitor de CMMS
mes=[0,0,0,0,0,0,0,0,0,0,0]
tw=[0,0,0,0,0,0,0,0,0,0,0]
t_ope=[0,0,0,0,0,0,0,0,0,0,0]
t_pa=[0,0,0,0,0,0,0,0,0,0,0]
idT=[0,0,0,0,0,0,0,0,0,0,0]
tT=[0,0,0,0,0,0,0,0,0,0,0]
mttr=[0,0,0,0,0,0,0,0,0,0,0]
mtbf=[0,0,0,0,0,0,0,0,0,0,0]

#REPORTE DE FALLAS MES
fail=[0,0,0,0,0,0,0,0,0,0,0]
date=[0,0,0,0,0,0,0,0,0,0,0]
est=[0,0,0,0,0,0,0,0,0,0,0] 
tr=[0,0,0,0,0,0,0,0,0,0,0]

tf=[0,0,0,0,0,0,0,0,0,0,0]

app = Flask(__name__)

#data_queue = Queue()
data_queue = queue.Queue()
pa_q=queue.Queue()
tca_q=queue.Queue()
tf_q=queue.Queue()
idalm_q=queue.Queue()
Andon_q=queue.Queue()
status_q=queue.Queue()

"""
#------------------Codigo para lectura a PLC en THREAD-------------------
fins = finsudp.FinsUDP(0 , 170)
for i in range(11):
           plcip= '192.168.10.' + ip[i]
           ret = fins.open(plcip, 9600)
           fins.set_destination(dst_net_addr=0, dst_node_num=ipst[i], dst_unit_addr=0)
           #words
           ret , pa[i] =fins.read_mem_area(datadef.DM_WORD, 10122,0,4, datadef.USHORT)
           ret , tca[i] =fins.read_mem_area(datadef.DM_WORD, 10122,0,4, datadef.USHORT)
           ret , tf[i] =fins.read_mem_area(datadef.DM_WORD, 10122,0,4, datadef.USHORT)
           ret , idalm[i] =fins.read_mem_area(datadef.DM_WORD, 10122,0,4, datadef.USHORT)
           ret , Andn[i] =fins.read_mem_area(datadef.DM_WORD, 10122,0,4, datadef.USHORT)
           #bits
           ret , status[i] =fins.read_mem_area(datadef.WR_BIT, 250,0,1, datadef.BIT)                  
        #----------------------------------------                      


def PLC_sample():
    
    while True:
        #Codigo para lectura a PLC en THREAD
        fins = finsudp.FinsUDP(0 , 170)
        for i in range(11):
           plcip= '192.168.10.' + ip[i]
           ret = fins.open(plcip, 9600)
           fins.set_destination(dst_net_addr=0, dst_node_num=ipst[i], dst_unit_addr=0)
           #words
           ret , pa[i] =fins.read_mem_area(datadef.DM_WORD, 310,0,1, datadef.USHORT)
           ret , tca[i] =fins.read_mem_area(datadef.DM_WORD, 290,0,1, datadef.USHORT)
           ret , tf[i] =fins.read_mem_area(datadef.DM_WORD, 292,0,1, datadef.USHORT)
           ret , idalm[i] =fins.read_mem_area(datadef.DM_WORD, 285,0,1, datadef.USHORT)
           ret , Andn[i] =fins.read_mem_area(datadef.DM_WORD, 300,0,1, datadef.USHORT)
           #bits
           ret , status[i] =fins.read_mem_area(datadef.WR_BIT, 282,0,1, datadef.BIT)                  
        #----------------------------------------                      
        #data_queue.put(pa) 
        pa_q.put(pa)
        tca_q.put(tca)
        tf_q.put(tf)
        idalm_q.put(idalm)
        Andon_q.put(Andon)
        status_q.put(status)
        
        time.sleep(5)

threadPLC = threading.Thread(target=PLC_sample, daemon=True)
threadPLC.start()
 #----------------------------------------------------------------------------------"""


#-------------------codigo para simulación---------------------------------------

workbook = load_workbook(ruta)
worksheet = workbook['Hoja1']

for i in range(11):
  pa[i] = worksheet.cell(row=i+2,column=2).value
  tca[i] = worksheet.cell(row=i+2,column=5).value
  tf[i] = worksheet.cell(row=i+2,column=8).value
  idalm[i] = worksheet.cell(row=i+2,column=21).value
  Andn[i] = worksheet.cell(row=i+2,column=12).value
  status[i] = worksheet.cell(row=i+2,column=20).value

workbook.close()

def monitor_fallas():
    while True:
        workbook = load_workbook(ruta)
        worksheet = workbook['Hoja1']
        for i in range(11):
          pa[i] = worksheet.cell(row=i+2,column=2).value
          tca[i] = worksheet.cell(row=i+2,column=5).value
          tf[i] = worksheet.cell(row=i+2,column=8).value
          idalm[i] = worksheet.cell(row=i+2,column=21).value
          Andn[i] = worksheet.cell(row=i+2,column=12).value
          status[i] = worksheet.cell(row=i+2,column=20).value
      
        workbook.close()           
        
        pa_q.put(pa)
        tca_q.put(tca)
        tf_q.put(tf)
        idalm_q.put(idalm)
        Andon_q.put(Andn)
        status_q.put(status)
        
        time.sleep(2)
           
thread = threading.Thread(target=monitor_fallas, daemon=True)
thread.start()
#thread.join()
#-----------------------------------------------------------------------------"""

def AndonRR1():
     
     st_and = Andon_q.get()
         
     for i in range(6):
       
       if st_and[i] == 1:
        mat1[i] =  0
        ld1[i] = 0
        mt1[i] = 0
        cal1[i]= 1
        
       if st_and[i] == 2:
        mat1[i] =  0
        ld1[i] = 0
        mt1[i] = 1
        cal1[i]= 0
        
       if st_and[i] == 4:
        mat1[i] =  0
        ld1[i] = 1
        mt1[i] = 0
        cal1[i]= 0
        
       if st_and[i] == 8:
        mat1[i] = 1
        ld1[i] = 0
        mt1[i] = 0
        cal1[i]= 0
       
       if st_and[i] != 1 and st_and[i] != 2 and st_and[i] != 4 and st_and[i] != 8:
        mat1[i] =  0
        ld1[i] = 0
        mt1[i] = 0
        cal1[i]= 0

       # Example data
     data1 = [
        {'st':' RR100A','Mat1': mat1[0], 'ld1': ld1[0], 'Mt1': mt1[0], 'Cal1': cal1[0]},
        {'st':' RR90A','Mat1': mat1[1], 'ld1': ld1[1], 'Mt1': mt1[1], 'Cal1': cal1[1]},
        {'st':' RR80A','Mat1': mat1[2], 'ld1': ld1[2], 'Mt1': mt1[2], 'Cal1': cal1[2]},
        {'st':' RR70A','Mat1': mat1[3], 'ld1': ld1[3], 'Mt1': mt1[3], 'Cal1': cal1[3]},
        {'st':' RR60A','Mat1': mat1[4], 'ld1': ld1[4], 'Mt1': mt1[4], 'Cal1': cal1[4]},
        {'st':' RR50A','Mat1': mat1[5], 'ld1': ld1[5], 'Mt1': mt1[5], 'Cal1': cal1[5]},

         ]
     return data1   

def AndonFR2():
      
     st_and = Andon_q.get() 
     
     for i in range(5):
       
       if st_and[i+6] == 1:
        mat2[i] =  0
        ld2[i] = 0
        mt2[i] = 0
        cal2[i]= 1
        
       if st_and[i+6] == 2:
        mat2[i] =  0
        ld2[i] = 0
        mt2[i] = 1
        cal2[i]= 0
        
       if st_and[i+6] == 4:
        mat2[i] =  0
        ld2[i] = 1
        mt2[i] = 0
        cal2[i]= 0
        
       if st_and[i+6] == 8:
        mat2[i] =  1
        ld2[i] = 0
        mt2[i] = 0
        cal2[i]= 0
        
       if st_and[i+6] != 1 and st_and[i+6] != 2 and st_and[i+6] != 4 and st_and[i+6] != 8:
        mat2[i] =  0
        ld2[i] = 0
        mt2[i] = 0
        cal2[i]= 0
      
       # Example data
     data2 = [
       {'st2':' FR80B','Mat2': mat2[0], 'ld2': ld2[0], 'Mt2': mt2[0], 'Cal2': cal2[0]},
       {'st2':' FR70B','Mat2': mat2[1], 'ld2': ld2[1], 'Mt2': mt2[1], 'Cal2': cal2[1]},
       {'st2':' FR60B','Mat2': mat2[2], 'ld2': ld2[2], 'Mt2': mt2[2], 'Cal2': cal2[2]},
       {'st2':' FR50B','Mat2': mat2[3], 'ld2': ld2[3], 'Mt2': mt2[3], 'Cal2': cal2[3]},
       {'st2':' FR60B','Mat2': mat2[4], 'ld2': ld2[4], 'Mt2': mt2[4], 'Cal2': cal2[4]}
       ]
     return data2   

def cmms1():
      
     workbook = load_workbook(ruta2)
     worksheet = workbook['CMMS']
     
     for i in range(11):
       mes[i] = worksheet.cell(row=i+2,column=2).value
       tw[i] = worksheet.cell(row=i+2,column=3).value
       t_ope[i] = worksheet.cell(row=i+2,column=4).value
       t_pa[i]= worksheet.cell(row=i+2,column=5).value
       idT[i] = worksheet.cell(row=i+2,column=6).value
       tT[i] = worksheet.cell(row=i+2,column=7).value
       mttr[i] = worksheet.cell(row=i+2,column=8).value
       mtbf[i]= worksheet.cell(row=i+2,column=9).value
       
     workbook.close()
     
       # Example data
     data_cmms = [
        {'st': 'RR100A', 'mes': mes[0], 'tw': tw[0], 't_ope': t_ope[0], 't_pa': t_pa[0], 'idT': idT[0], 'tT': tT[0], 'mttr': mttr[0], 'mtbf': mtbf[0]},
        {'st': 'RR90A', 'mes': mes[1], 'tw': tw[1], 't_ope': t_ope[1], 't_pa': t_pa[1], 'idT': idT[1], 'tT': tT[1], 'mttr': mttr[1], 'mtbf': mtbf[1]},
        {'st': 'RR80A', 'mes': mes[2], 'tw': tw[2], 't_ope': t_ope[2], 't_pa': t_pa[2], 'idT': idT[2], 'tT': tT[2], 'mttr': mttr[2], 'mtbf': mtbf[2]},
        {'st': 'RR70A', 'mes': mes[3], 'tw': tw[3], 't_ope': t_ope[3], 't_pa': t_pa[3], 'idT': idT[3], 'tT': tT[3], 'mttr': mttr[3], 'mtbf': mtbf[3]},
        {'st': 'RR60A', 'mes': mes[4], 'tw': tw[4], 't_ope': t_ope[4], 't_pa': t_pa[4], 'idT': idT[4], 'tT': tT[4], 'mttr': mttr[4], 'mtbf': mtbf[4]},
        {'st': 'RR50A', 'mes': mes[5], 'tw': tw[5], 't_ope': t_ope[5], 't_pa': t_pa[5], 'idT': idT[5], 'tT': tT[5], 'mttr': mttr[5], 'mtbf': mtbf[5]},
        {'st': 'FR80A', 'mes': mes[6], 'tw': tw[6], 't_ope': t_ope[6], 't_pa': t_pa[6], 'idT': idT[6], 'tT': tT[6], 'mttr': mttr[6], 'mtbf': mtbf[6]},
        {'st': 'FR70A', 'mes': mes[7], 'tw': tw[7], 't_ope': t_ope[7], 't_pa': t_pa[7], 'idT': idT[7], 'tT': tT[7], 'mttr': mttr[7], 'mtbf': mtbf[7]},
        {'st': 'FR60A', 'mes': mes[8], 'tw': tw[8], 't_ope': t_ope[8], 't_pa': t_pa[8], 'idT': idT[8], 'tT': tT[8], 'mttr': mttr[8], 'mtbf': mtbf[8]},
        {'st': 'FR50A', 'mes': mes[9], 'tw': tw[9], 't_ope': t_ope[9], 't_pa': t_pa[9], 'idT': idT[9], 'tT': tT[9], 'mttr': mttr[9], 'mtbf': mtbf[9]},
        {'st': 'FR60A', 'mes': mes[10], 'tw': tw[10], 't_ope': t_ope[10], 't_pa': t_pa[10], 'idT': idT[10], 'tT': tT[10], 'mttr': mttr[10], 'mtbf': mtbf[10]}


         ]
     return data_cmms

def report():

     workbook = load_workbook(ruta2)
     worksheet = workbook['fallas']
     for i in range(11):
        fail[i] = worksheet.cell(row=i+2,column=2).value
        date[i] = worksheet.cell(row=i+2,column=3).value
        est[i] = worksheet.cell(row=i+2,column=4).value
        tr[i]= worksheet.cell(row=i+2,column=5).value
        
     workbook.close()

       # Example data
     datafa = [
       {'fail':fail[0],'date': date[0], 'est': est[0], 'tr': tr[0]},
       {'fail':fail[1],'date': date[1], 'est': est[1], 'tr': tr[1]},
       {'fail':fail[2],'date': date[2], 'est': est[2], 'tr': tr[2]},
       {'fail':fail[3],'date': date[3], 'est': est[3], 'tr': tr[3]},
       {'fail':fail[4],'date': date[4], 'est': est[4], 'tr': tr[4]},
       {'fail':fail[5],'date': date[5], 'est': est[5], 'tr': tr[5]},
       {'fail':fail[6],'date': date[6], 'est': est[6], 'tr': tr[6]},
       {'fail':fail[7],'date': date[7], 'est': est[7], 'tr': tr[7]},
       {'fail':fail[8],'date': date[8], 'est': est[8], 'tr': tr[8]},
       {'fail':fail[9],'date': date[9], 'est': est[9], 'tr': tr[9]},
       {'fail':fail[10],'date': date[10], 'est': est[10], 'tr': tr[10]}

         ]
     return datafa   

def failures():
     
     stat = status_q.get()
     idfail = idalm_q.get()
             
     datafail = [
          {'stF': 'RR100A', 'NOW': stat[0], 'ID': idfail[0]},
          {'stF': 'RR90A', 'NOW': stat[1], 'ID': idfail[1]},
          {'stF': 'RR80A', 'NOW': stat[2], 'ID': idfail[2]},
          {'stF': 'RR70A', 'NOW': stat[3], 'ID': idfail[3]},
          {'stF': 'RR60A', 'NOW': stat[4], 'ID': idfail[4]},
          {'stF': 'RR50A', 'NOW': stat[5], 'ID': idfail[5]},
          {'stF': 'FR80A', 'NOW': stat[6], 'ID': idfail[6]},
          {'stF': 'FR70A', 'NOW': stat[7], 'ID': idfail[7]},
          {'stF': 'FR60A', 'NOW': stat[8], 'ID': idfail[8]},
          {'stF': 'FR50A', 'NOW': stat[9], 'ID': idfail[9]},
          {'stF': 'FR60A', 'NOW': stat[10], 'ID': idfail[10]}
                 ]      
     return datafail   



@app.route('/')
def principal ():    
    # Replace with logic to retrieve data from your PLC or data source
    return render_template('index.html')
    
@app.route('/data')
def get_data():
    #valores producción
    
    #calculo de turno actual
    c_time = datetime.datetime.now()
    c_hour = c_time.hour
    c_min = c_time.minute    
    c_seg = c_time.second
    
    if c_hour >= 8 and c_hour < 20:
      shift = 1
      
    else:
      shift = 2 
        
    #calculo de tiempo de paro programado
    if c_hour >= 8 and c_hour < 10 and  shift == 1:
      tpp = 0
    
    if c_hour >= 10 and c_hour < 12 and  shift == 1:
      tpp = 10
      
    if c_hour >= 12 and c_hour < 15 and shift == 1:
      tpp = 50
      
    if c_hour >= 15 and c_hour < 20 and shift == 1:
      tpp = 60
      
    if c_hour >= 22 and c_hour > 1 and  shift == 2:
      tpp = 10
      
    if c_hour >= 0  and c_hour < 3 and shift == 2:
      tpp = 50
      
    if c_hour >= 3 and c_hour < 8 and shift == 2:
      tpp = 60
      
    #calculo de tiempo total de operacion  
    if shift == 1:
      hora_t = c_hour - 8  
      tot=(hora_t*3600)+(c_min*60)+c_seg
    
    if shift == 2 and c_hour >0 and c_hour < 8:
      hora_t = c_hour  
      tot=(hora_t*3600)+(c_min*60)+c_seg
      
    if shift == 2 and c_hour >20 and c_hour <= 23:
      hora_t = c_hour-20  
      tot=(hora_t*3600)+(c_min*60)+c_seg
      
    workbook = load_workbook(ruta)
    
    worksheet = workbook['Hoja1']
    ac=pa_q.get()
    tep=tca_q.get()
    tp_m=tf_q.get()
    
       
    for i in range(11):
      
       #produccion actual
       #ac[i] = worksheet.cell(row=i+2,column=2).value
       #ac=pa_q.get()
       
       #plan de produccion
       pl[i] = worksheet.cell(row=i+2,column=3).value
       #diferencia Plan -actual
       df[i] = pl[i] - ac[i]       
       #Tiempo Total. Tiempo total desde inicio de turno
       tt[i]=round(tot/60,2)
       #tiempo efectivo maquina. Tiempo maquina de operación automatica
       #tep = worksheet.cell(row=i+2,column=5).value
       te[i] = round(tep[i]/60,2)
       #tiempo hombre
       th[i] = worksheet.cell(row=i+2,column=6).value
       thpp = th[i] * ac[i]
       thp[i] = round(thpp/60,2)
       #tiempo de paro planeado
       pp[i]=tpp 
       #tiempo de paros por anomalias
       #tp[i] = worksheet.cell(row=i+2,column=8).value     
       tp[i] = round(tp_m[i]/60,2)
       #cálculo de tiempo sin operación 
       tm[i] = (tot/60) - te[i]
       #calculo de tiempo muerto
       tmp= tm[i] - tp[i] - thp[i] - pp[i]
       tmf[i]=round(tmp,2)
       #calculo de tiempo ciclo proceso
       tcp = ((tot/60) - tpp) / ac[i]
       tc[i]= round (tcp,2)
       #calculo de OEE
       tpo = (tot/60) - tpp 
       to= tpo -tmf[i]    
       efp = (to / tpo) * 100
       ef[i] = round (efp, 2)
       
    workbook.close()
    """st: estación - pa: producción actual"""  
    
    data = [       
        {'st':' RR100A','pa': ac[0], 'plan': pl[0], 'df': df[0], 'tt': tt[0], 'te': te[0], 'thp': thp[0], 'pp': pp[0], 'tp': tp[0], 'tmf': tmf[0], 'tc': tc[0], 'ef': ef[0]},
        {'st':' RR90A','pa': ac[1], 'plan': pl[1], 'df': df[1], 'tt': tt[1], 'te': te[1], 'thp': thp[1], 'pp': pp[1], 'tp': tp[1], 'tmf': tmf[1], 'tc': tc[1], 'ef': ef[1]},
        {'st':' RR80A','pa': ac[2], 'plan': pl[2], 'df': df[2], 'tt': tt[2], 'te': te[2], 'thp': thp[2], 'pp': pp[2], 'tp': tp[2], 'tmf': tmf[2], 'tc': tc[2], 'ef': ef[2]},
        {'st':' RR70A','pa': ac[3], 'plan': pl[3], 'df': df[3], 'tt': tt[3], 'te': te[3], 'thp': thp[3], 'pp': pp[3], 'tp': tp[3], 'tmf': tmf[3], 'tc': tc[3], 'ef': ef[3]},
        {'st':' RR60A','pa': ac[4], 'plan': pl[4], 'df': df[4], 'tt': tt[4], 'te': te[4], 'thp': thp[4], 'pp': pp[4], 'tp': tp[4], 'tmf': tmf[4], 'tc': tc[4], 'ef': ef[4]},
        {'st':' RR50A','pa': ac[5], 'plan': pl[5], 'df': df[5], 'tt': tt[5], 'te': te[5], 'thp': thp[5], 'pp': pp[5], 'tp': tp[5], 'tmf': tmf[5], 'tc': tc[5], 'ef': ef[5]},
        {'st':' FR80A','pa': ac[6], 'plan': pl[6], 'df': df[6], 'tt': tt[6], 'te': te[6], 'thp': thp[6], 'pp': pp[6], 'tp': tp[6], 'tmf': tmf[6], 'tc': tc[6], 'ef': ef[6]},
        {'st':' FR70A','pa': ac[7], 'plan': pl[7], 'df': df[7], 'tt': tt[7], 'te': te[7], 'thp': thp[7], 'pp': pp[7], 'tp': tp[7], 'tmf': tmf[7], 'tc': tc[7], 'ef': ef[7]},
        {'st':' FR60A','pa': ac[8], 'plan': pl[8], 'df': df[8], 'tt': tt[8], 'te': te[8], 'thp': thp[8], 'pp': pp[8], 'tp': tp[8], 'tmf': tmf[8], 'tc': tc[8], 'ef': ef[8]},
        {'st':' FR50A','pa': ac[9], 'plan': pl[9], 'df': df[9], 'tt': tt[9], 'te': te[9], 'thp': thp[9], 'pp': pp[9], 'tp': tp[9], 'tmf': tmf[9], 'tc': tc[9], 'ef': ef[9]},
        {'st':' FR60A','pa': ac[10], 'plan': pl[10], 'df': df[10], 'tt': tt[10], 'te': te[10], 'thp': thp[10], 'pp': pp[10], 'tp': tp[10], 'tmf': tmf[10], 'tc': tc[10], 'ef': ef[10]}

    ]
    
    return jsonify(data=data)


@app.route('/dataAndon')
def get_dataAndon():
    
    dataAndon = AndonRR1()  
    
    return jsonify(dataAdn=dataAndon)


@app.route('/Reporte')
def Reporte():
    return render_template('Reporte.html')


@app.route('/Andon')
def Andon (): 
    return render_template('Andon.html',)   

@app.route('/cmms')
def cmms():

    return render_template('cmms.html' )

@app.route('/andon2')
def andon2():
    signal = 0
    data1 = AndonRR1()
    data2= AndonFR2()    
    return render_template('andon2.html', data=data1, sig=signal, dataB=data2 )


@app.route('/update_data')
def update_data():
       
  data1=AndonRR1()
     
  return jsonify(data1)


@app.route('/update_data2')
def update_data2():
    
  data2=AndonFR2()      
    
  return jsonify(data2)      

@app.route('/update_datacmms')
def update_datacmms():
    D_cmms = cmms1()      
    return jsonify(data =D_cmms)
  
@app.route('/update_datafa')
def update_datafa():
    rep = report()      
    return jsonify(data = rep)


@app.route('/fallas')
def fallas ():
    Tfallas= failures()    
  
    return render_template('fallas.html', fallas=Tfallas)

    
@app.route('/update_data3')
def update_data3():  
    Tfallas= failures()
         
    return jsonify(Tfallas)


if __name__ == '__main__':
    app.run(debug=True)